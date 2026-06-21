# -*- coding: utf-8 -*-
"""
routes/export_routes.py — Xuất báo cáo phân tích cổ phiếu ra file Excel.

Tạo file .xlsx với 4 sheet:
  1. Thông Tin       — thông tin chung (tên, ngành, sàn, giá, ...)
  2. Giá Lịch Sử    — bảng OHLCV
  3. Phân Tích KT   — chỉ báo kỹ thuật (MA, RSI, MACD, Bollinger)
  4. Phân Tích CB   — chỉ số cơ bản (ROE, ROA, EPS, PE, PB, DE) + chấm điểm
"""

import io
from datetime import datetime

from flask import Blueprint, send_file, jsonify

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from modules import module1_thudulieu as module1
from modules import module2_kythuat   as module2
from modules import module3_coban     as module3

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

export_bp = Blueprint("export", __name__)


# ---------------------------------------------------------------------------
# Bảng màu
# ---------------------------------------------------------------------------

_CLR_HEADER_DARK  = "1A1D2E"   # header chính — nền tối navy
_CLR_HEADER_MID   = "252A3D"   # header phụ — nền tối nhạt hơn
_CLR_ACCENT_BLUE  = "3B82F6"   # xanh dương
_CLR_ACCENT_GREEN = "22C55E"   # xanh lá
_CLR_ACCENT_RED   = "EF4444"   # đỏ
_CLR_ACCENT_ORG   = "F97316"   # cam
_CLR_ACCENT_YELL  = "EAB308"   # vàng
_CLR_ROW_EVEN     = "F1F5F9"   # hàng chẵn xám nhạt
_CLR_ROW_ODD      = "FFFFFF"   # hàng lẻ trắng
_CLR_TITLE_TEXT   = "FFFFFF"   # chữ trắng trên nền tối
_CLR_LABEL        = "374151"   # chữ tối cho nhãn
_CLR_VALUE        = "111827"   # chữ đậm cho giá trị


def _side(style="thin", color="D1D5DB"):
    return Side(style=style, color=color)

def _border_thin():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium():
    s = _side("medium", "9CA3AF")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color="111827", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _right():
    return Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Helper: tô màu ô header
# ---------------------------------------------------------------------------

def _apply_header(cell, text, color=_CLR_HEADER_DARK, size=11):
    cell.value = text
    cell.font = _font(bold=True, size=size, color=_CLR_TITLE_TEXT)
    cell.fill = _fill(color)
    cell.alignment = _center()
    cell.border = _border_thin()


def _apply_label(cell, text):
    cell.value = text
    cell.font = _font(bold=True, size=11, color=_CLR_LABEL)
    cell.fill = _fill(_CLR_ROW_EVEN)
    cell.alignment = _left()
    cell.border = _border_thin()


def _apply_value(cell, value, number_format=None, bold=False):
    cell.value = value
    cell.font = _font(bold=bold, size=11, color=_CLR_VALUE)
    cell.alignment = _right()
    cell.border = _border_thin()
    if number_format:
        cell.number_format = number_format


def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Sheet 1 — Thông tin chung
# ---------------------------------------------------------------------------

def _sheet_thong_tin(wb, info, ma_cp: str):
    ws = wb.create_sheet("Thông Tin")
    ws.sheet_view.showGridLines = False

    # Tiêu đề lớn
    ws.merge_cells("A1:D1")
    cell_title = ws["A1"]
    cell_title.value = f"THÔNG TIN CỔ PHIẾU  {ma_cp}"
    cell_title.font = _font(bold=True, size=16, color=_CLR_TITLE_TEXT)
    cell_title.fill = _fill(_CLR_HEADER_DARK)
    cell_title.alignment = _center()
    ws.row_dimensions[1].height = 36

    # Sub-header
    ws.merge_cells("A2:D2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ts_cell.font = _font(italic=True, size=10, color=_CLR_TITLE_TEXT)
    ts_cell.fill = _fill(_CLR_HEADER_MID)
    ts_cell.alignment = _center()
    ws.row_dimensions[2].height = 20

    # Dữ liệu
    rows = [
        ("Mã cổ phiếu",       info.get("ma", ma_cp)),
        ("Tên công ty",        info.get("ten_cong_ty", "—")),
        ("Ngành",              info.get("nganh", "—")),
        ("Sàn giao dịch",     info.get("san", "—")),
        ("Giá hiện tại (nghìn đ)", info.get("gia_hien_tai")),
        ("Thay đổi (%)",      info.get("thay_doi_phan_tram")),
        ("Khối lượng GD",     info.get("khoi_luong")),
        ("Vốn hóa (tỷ đ)",   info.get("von_hoa")),
    ]

    for i, (label, val) in enumerate(rows):
        r = i + 4
        _apply_label(ws[f"A{r}"], label)
        ws.merge_cells(f"B{r}:D{r}")
        vtype = None
        if isinstance(val, float) and label.startswith("Giá"):
            vtype = '#,##0.00'
        elif isinstance(val, float) and "%" in label:
            vtype = '#,##0.00'
        elif isinstance(val, (int, float)):
            vtype = '#,##0'
        _apply_value(ws[f"B{r}"], val, vtype)

    # Widths
    _set_col_width(ws, "A", 28)
    _set_col_width(ws, "B", 20)
    _set_col_width(ws, "C", 20)
    _set_col_width(ws, "D", 20)


# ---------------------------------------------------------------------------
# Sheet 2 — Giá lịch sử OHLCV
# ---------------------------------------------------------------------------

def _sheet_gia_lich_su(wb, df_gia):
    ws = wb.create_sheet("Giá Lịch Sử")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Tiêu đề
    ws.merge_cells("A1:F1")
    cell_title = ws["A1"]
    cell_title.value = "GIÁ LỊCH SỬ (OHLCV)"
    cell_title.font = _font(bold=True, size=14, color=_CLR_TITLE_TEXT)
    cell_title.fill = _fill(_CLR_ACCENT_BLUE)
    cell_title.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Header cột
    headers = ["Ngày", "Mở cửa", "Cao nhất", "Thấp nhất", "Đóng cửa", "Khối lượng"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=j)
        _apply_header(c, h, _CLR_HEADER_MID)
    ws.row_dimensions[2].height = 22

    # Dữ liệu
    for i, (_, row) in enumerate(df_gia.iterrows()):
        r = i + 3
        bg = _CLR_ROW_EVEN if i % 2 == 0 else _CLR_ROW_ODD
        fill = _fill(bg)
        bdr  = _border_thin()

        date_val = row["date"].date() if hasattr(row["date"], "date") else row["date"]

        cells_data = [
            (date_val, "DD/MM/YYYY", True),
            (float(row["open"]),   "#,##0.00", False),
            (float(row["high"]),   "#,##0.00", False),
            (float(row["low"]),    "#,##0.00", False),
            (float(row["close"]), "#,##0.00", False),
            (int(row["volume"]),   "#,##0",    False),
        ]
        for j, (val, fmt, center) in enumerate(cells_data, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.number_format = fmt
            c.font = _font(size=10)
            c.fill = fill
            c.border = bdr
            c.alignment = _center() if center else _right()

    # Widths
    widths = [14, 14, 14, 14, 14, 16]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------------------
# Sheet 3 — Phân tích kỹ thuật
# ---------------------------------------------------------------------------

def _sheet_ky_thuat(wb, tech: dict):
    ws = wb.create_sheet("Phân Tích KT")
    ws.sheet_view.showGridLines = False

    # Tiêu đề
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "PHÂN TÍCH KỸ THUẬT"
    t.font = _font(bold=True, size=14, color=_CLR_TITLE_TEXT)
    t.fill = _fill(_CLR_ACCENT_GREEN)
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Tín hiệu tổng hợp
    tin_hieu = tech.get("tin_hieu", "—")
    tin_hieu_color = {
        "MUA MẠNH": _CLR_ACCENT_GREEN,
        "MUA":       "16A34A",
        "GIỮ":       _CLR_ACCENT_YELL,
        "BÁN":       _CLR_ACCENT_RED,
    }.get(tin_hieu, _CLR_HEADER_MID)

    ws.merge_cells("A2:D2")
    sig = ws["A2"]
    sig.value = f"Tín hiệu: {tin_hieu}  |  Số tín hiệu mua: {tech.get('so_tin_hieu_mua', '—')}/4"
    sig.font = _font(bold=True, size=12, color=_CLR_TITLE_TEXT)
    sig.fill = _fill(tin_hieu_color)
    sig.alignment = _center()
    ws.row_dimensions[2].height = 24

    # Giải thích
    ws.merge_cells("A3:D3")
    exp = ws["A3"]
    exp.value = tech.get("giai_thich", "")
    exp.font = _font(italic=True, size=10, color="4B5563")
    exp.fill = _fill("F8FAFC")
    exp.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    exp.border = _border_thin()
    ws.row_dimensions[3].height = 40

    # Header bảng chỉ báo
    headers_kt = ["Chỉ báo", "Giá trị", "Đơn vị", "Mô tả"]
    for j, h in enumerate(headers_kt, 1):
        c = ws.cell(row=5, column=j)
        _apply_header(c, h, _CLR_HEADER_DARK)
    ws.row_dimensions[5].height = 22

    def pct_rsi(v):
        if v is None: return "—"
        if v >= 70: return "Quá mua"
        if v <= 30: return "Quá bán"
        return "Trung tính"

    rows_kt = [
        ("RSI (14)",          tech.get("rsi"),           "",       pct_rsi(tech.get("rsi"))),
        ("MACD",              tech.get("macd"),           "",       "MACD Line"),
        ("Signal",            tech.get("signal"),         "",       "Signal Line (EMA9)"),
        ("Histogram",         tech.get("histogram"),      "",       "MACD - Signal"),
        ("MA20",              tech.get("ma20"),           "nghìn đ", "Trung bình 20 phiên"),
        ("MA50",              tech.get("ma50"),           "nghìn đ", "Trung bình 50 phiên"),
        ("MA200",             tech.get("ma200"),          "nghìn đ", "Trung bình 200 phiên"),
        ("Bollinger Upper",   tech.get("bollinger_upper"),  "nghìn đ", "Dải trên"),
        ("Bollinger Middle",  tech.get("bollinger_middle"), "nghìn đ", "Dải giữa (SMA20)"),
        ("Bollinger Lower",   tech.get("bollinger_lower"),  "nghìn đ", "Dải dưới"),
    ]

    for i, (label, val, unit, desc) in enumerate(rows_kt):
        r = i + 6
        bg = _CLR_ROW_EVEN if i % 2 == 0 else _CLR_ROW_ODD
        fill = _fill(bg)
        bdr  = _border_thin()

        ws.cell(row=r, column=1, value=label).font   = _font(bold=True, size=10)
        ws.cell(row=r, column=1).fill   = fill
        ws.cell(row=r, column=1).border = bdr
        ws.cell(row=r, column=1).alignment = _left()

        val_display = round(val, 4) if isinstance(val, float) else (val if val is not None else "—")
        ws.cell(row=r, column=2, value=val_display).font   = _font(size=10)
        ws.cell(row=r, column=2).fill   = fill
        ws.cell(row=r, column=2).border = bdr
        ws.cell(row=r, column=2).alignment = _right()
        if isinstance(val, float):
            ws.cell(row=r, column=2).number_format = "#,##0.00"

        ws.cell(row=r, column=3, value=unit).font   = _font(italic=True, size=10, color="6B7280")
        ws.cell(row=r, column=3).fill   = fill
        ws.cell(row=r, column=3).border = bdr
        ws.cell(row=r, column=3).alignment = _center()

        ws.cell(row=r, column=4, value=desc).font   = _font(size=10, color="6B7280")
        ws.cell(row=r, column=4).fill   = fill
        ws.cell(row=r, column=4).border = bdr
        ws.cell(row=r, column=4).alignment = _left()

    _set_col_width(ws, "A", 22)
    _set_col_width(ws, "B", 16)
    _set_col_width(ws, "C", 14)
    _set_col_width(ws, "D", 36)


# ---------------------------------------------------------------------------
# Sheet 4 — Phân tích cơ bản
# ---------------------------------------------------------------------------

def _sheet_co_ban(wb, fund: dict):
    ws = wb.create_sheet("Phân Tích CB")
    ws.sheet_view.showGridLines = False

    chi_so    = fund.get("chi_so",    {})
    cham_diem = fund.get("cham_diem", {})
    tong      = cham_diem.get("tong",    0)
    phan_loai = cham_diem.get("phan_loai", "—")

    # Tiêu đề
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "PHÂN TÍCH CƠ BẢN"
    t.font = _font(bold=True, size=14, color=_CLR_TITLE_TEXT)
    t.fill = _fill(_CLR_ACCENT_ORG)
    t.alignment = _center()
    ws.row_dimensions[1].height = 30

    # Điểm tổng
    score_color = (
        _CLR_ACCENT_GREEN if phan_loai == "TỐT"
        else _CLR_ACCENT_ORG if phan_loai == "KHÁ"
        else _CLR_ACCENT_RED
    )
    ws.merge_cells("A2:E2")
    sc = ws["A2"]
    sc.value = f"Điểm tổng: {tong}/12  —  Xếp loại: {phan_loai}"
    sc.font = _font(bold=True, size=13, color=_CLR_TITLE_TEXT)
    sc.fill = _fill(score_color)
    sc.alignment = _center()
    ws.row_dimensions[2].height = 26

    # Header bảng
    headers_cb = ["Chỉ số", "Giá trị", "Đơn vị", "Điểm", "Tiêu chí"]
    for j, h in enumerate(headers_cb, 1):
        c = ws.cell(row=4, column=j)
        _apply_header(c, h, _CLR_HEADER_DARK)
    ws.row_dimensions[4].height = 22

    meta = {
        "ROE": ("%",   "≥20%: 2đ | 15–20%: 1đ | <15%: 0đ"),
        "ROA": ("%",   "≥10%: 2đ | 5–10%: 1đ  | <5%:  0đ"),
        "EPS": ("₫/CP","≥5000: 2đ | 2000–5000: 1đ | <2000: 0đ"),
        "PE":  ("x",   "8–15x: 2đ | 15–20x: 1đ | ngoài: 0đ"),
        "PB":  ("x",   "<1.5x: 2đ | 1.5–2.5x: 1đ | >2.5x: 0đ"),
        "DE":  ("",    "<1: 2đ | 1–2: 1đ | >2: 0đ"),
    }

    for i, (key, (unit, tieu_chi)) in enumerate(meta.items()):
        r = i + 5
        bg = _CLR_ROW_EVEN if i % 2 == 0 else _CLR_ROW_ODD
        fill = _fill(bg)
        bdr  = _border_thin()

        val   = chi_so.get(key)
        score = cham_diem.get(key, 0)

        # Tô màu theo điểm
        score_bg = (
            "DCFCE7" if score == 2
            else "FEF9C3" if score == 1
            else "FEE2E2"
        )

        ws.cell(row=r, column=1, value=key).font        = _font(bold=True, size=11)
        ws.cell(row=r, column=1).fill   = fill
        ws.cell(row=r, column=1).border = bdr
        ws.cell(row=r, column=1).alignment = _center()

        val_disp = (
            round(val, 2) if isinstance(val, float) else
            int(val)      if isinstance(val, int)   else
            (val if val is not None else "—")
        )
        val_cell = ws.cell(row=r, column=2, value=val_disp)
        val_cell.font   = _font(bold=True, size=11)
        val_cell.fill   = fill
        val_cell.border = bdr
        val_cell.alignment = _right()
        if isinstance(val, float):
            val_cell.number_format = "#,##0.00"
        elif isinstance(val, int):
            val_cell.number_format = "#,##0"

        ws.cell(row=r, column=3, value=unit).font   = _font(italic=True, size=10, color="6B7280")
        ws.cell(row=r, column=3).fill   = fill
        ws.cell(row=r, column=3).border = bdr
        ws.cell(row=r, column=3).alignment = _center()

        score_cell = ws.cell(row=r, column=4, value=f"{score}/2")
        score_cell.font      = _font(bold=True, size=11)
        score_cell.fill      = _fill(score_bg)
        score_cell.border    = bdr
        score_cell.alignment = _center()

        ws.cell(row=r, column=5, value=tieu_chi).font   = _font(size=10, color="6B7280")
        ws.cell(row=r, column=5).fill   = fill
        ws.cell(row=r, column=5).border = bdr
        ws.cell(row=r, column=5).alignment = _left()

    # Tổng điểm row
    r_total = 5 + len(meta)
    ws.merge_cells(f"A{r_total}:C{r_total}")
    tot = ws[f"A{r_total}"]
    tot.value = "TỔNG ĐIỂM"
    tot.font = _font(bold=True, size=12, color=_CLR_TITLE_TEXT)
    tot.fill = _fill(_CLR_HEADER_DARK)
    tot.alignment = _center()
    tot.border = _border_medium()

    ws.cell(row=r_total, column=4, value=f"{tong}/12").font  = _font(bold=True, size=12, color=_CLR_TITLE_TEXT)
    ws.cell(row=r_total, column=4).fill   = _fill(score_color)
    ws.cell(row=r_total, column=4).border = _border_medium()
    ws.cell(row=r_total, column=4).alignment = _center()

    ws.cell(row=r_total, column=5, value=phan_loai).font  = _font(bold=True, size=12, color=_CLR_TITLE_TEXT)
    ws.cell(row=r_total, column=5).fill   = _fill(score_color)
    ws.cell(row=r_total, column=5).border = _border_medium()
    ws.cell(row=r_total, column=5).alignment = _center()

    ws.row_dimensions[r_total].height = 26

    _set_col_width(ws, "A", 10)
    _set_col_width(ws, "B", 16)
    _set_col_width(ws, "C", 10)
    _set_col_width(ws, "D", 10)
    _set_col_width(ws, "E", 42)


# ---------------------------------------------------------------------------
# Sheet tổng hợp (đầu tiên)
# ---------------------------------------------------------------------------

def _sheet_tong_hop(wb, info, tech, fund, ma_cp: str):
    ws = wb.create_sheet("Tổng Hợp", 0)   # chèn vào đầu
    ws.sheet_view.showGridLines = False

    chi_so    = fund.get("chi_so",    {})
    cham_diem = fund.get("cham_diem", {})
    tong      = cham_diem.get("tong", 0)
    phan_loai = cham_diem.get("phan_loai", "—")
    tin_hieu  = tech.get("tin_hieu", "—")

    # ---- Tiêu đề ----
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"BÁO CÁO PHÂN TÍCH CỔ PHIẾU  {ma_cp}"
    t.font = _font(bold=True, size=18, color=_CLR_TITLE_TEXT)
    t.fill = _fill(_CLR_HEADER_DARK)
    t.alignment = _center()
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = (
        f"{info.get('ten_cong_ty','—')}  |  "
        f"Sàn: {info.get('san','—')}  |  "
        f"Ngành: {info.get('nganh','—')}  |  "
        f"Xuất lúc: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    sub.font = _font(italic=True, size=10, color=_CLR_TITLE_TEXT)
    sub.fill = _fill(_CLR_HEADER_MID)
    sub.alignment = _center()
    ws.row_dimensions[2].height = 20

    # ---- Khung giá ----
    ws.merge_cells("A4:B4")
    _apply_header(ws["A4"], "Giá hiện tại", _CLR_ACCENT_BLUE, 12)
    ws.merge_cells("C4:D4")
    _apply_header(ws["C4"], "Thay đổi (%)", _CLR_ACCENT_BLUE, 12)
    ws.merge_cells("E4:F4")
    _apply_header(ws["E4"], "Tín hiệu KT", _CLR_ACCENT_GREEN, 12)

    ws.merge_cells("A5:B5")
    gp = ws["A5"]
    gp.value = info.get("gia_hien_tai")
    gp.number_format = "#,##0.00"
    gp.font = _font(bold=True, size=20, color=_CLR_ACCENT_BLUE)
    gp.fill = _fill("EFF6FF")
    gp.alignment = _center()
    gp.border = _border_thin()

    ws.merge_cells("C5:D5")
    pct = info.get("thay_doi_phan_tram", 0) or 0
    cp = ws["C5"]
    cp.value = pct
    cp.number_format = '+#,##0.00;-#,##0.00;0.00'
    cp.font = _font(bold=True, size=18,
                    color=_CLR_ACCENT_GREEN if pct > 0 else _CLR_ACCENT_RED)
    cp.fill = _fill("F0FDF4" if pct > 0 else "FEF2F2")
    cp.alignment = _center()
    cp.border = _border_thin()

    sig_color = {
        "MUA MẠNH": _CLR_ACCENT_GREEN,
        "MUA":       "16A34A",
        "GIỮ":       _CLR_ACCENT_YELL,
        "BÁN":       _CLR_ACCENT_RED,
    }.get(tin_hieu, _CLR_HEADER_MID)

    ws.merge_cells("E5:F5")
    sp = ws["E5"]
    sp.value = tin_hieu
    sp.font = _font(bold=True, size=16, color=_CLR_TITLE_TEXT)
    sp.fill = _fill(sig_color)
    sp.alignment = _center()
    sp.border = _border_thin()

    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 44

    # ---- Bảng tóm tắt chỉ số ----
    ws.merge_cells("A7:F7")
    sh = ws["A7"]
    sh.value = "TÓM TẮT CÁC CHỈ SỐ"
    sh.font = _font(bold=True, size=12, color=_CLR_TITLE_TEXT)
    sh.fill = _fill(_CLR_ACCENT_ORG)
    sh.alignment = _center()
    ws.row_dimensions[7].height = 24

    headers_sum = ["Chỉ số", "Giá trị", "Đvt", "Điểm CB", "Chỉ báo KT", "Giá trị KT"]
    for j, h in enumerate(headers_sum, 1):
        c = ws.cell(row=8, column=j)
        _apply_header(c, h, _CLR_HEADER_DARK)
    ws.row_dimensions[8].height = 22

    rows_sum = [
        ("ROE",  chi_so.get("ROE"), "%",        cham_diem.get("ROE"), "MA20",  tech.get("ma20")),
        ("ROA",  chi_so.get("ROA"), "%",        cham_diem.get("ROA"), "MA50",  tech.get("ma50")),
        ("EPS",  chi_so.get("EPS"), "₫/CP",     cham_diem.get("EPS"), "MA200", tech.get("ma200")),
        ("PE",   chi_so.get("PE"),  "x",        cham_diem.get("PE"),  "RSI",   tech.get("rsi")),
        ("PB",   chi_so.get("PB"),  "x",        cham_diem.get("PB"),  "MACD",  tech.get("macd")),
        ("DE",   chi_so.get("DE"),  "",         cham_diem.get("DE"),  "Sig",   tech.get("signal")),
    ]

    for i, (k, v, u, sc, kt_name, kt_val) in enumerate(rows_sum):
        r = i + 9
        bg = _CLR_ROW_EVEN if i % 2 == 0 else _CLR_ROW_ODD
        fill = _fill(bg)
        bdr  = _border_thin()

        score_bg = (
            "DCFCE7" if (sc or 0) == 2
            else "FEF9C3" if (sc or 0) == 1
            else "FEE2E2"
        )

        ws.cell(r, 1, k).font        = _font(bold=True, size=11)
        ws.cell(r, 1).fill   = fill; ws.cell(r, 1).border = bdr
        ws.cell(r, 1).alignment = _center()

        v_disp = round(v, 2) if isinstance(v, float) else (v if v is not None else "—")
        vc = ws.cell(r, 2, v_disp)
        vc.font = _font(size=11)
        vc.fill = fill; vc.border = bdr; vc.alignment = _right()
        if isinstance(v, float): vc.number_format = "#,##0.00"
        elif isinstance(v, int): vc.number_format = "#,##0"

        ws.cell(r, 3, u).font   = _font(italic=True, size=10, color="6B7280")
        ws.cell(r, 3).fill = fill; ws.cell(r, 3).border = bdr
        ws.cell(r, 3).alignment = _center()

        sc_cell = ws.cell(r, 4, f"{sc or 0}/2")
        sc_cell.font = _font(bold=True, size=11)
        sc_cell.fill = _fill(score_bg); sc_cell.border = bdr
        sc_cell.alignment = _center()

        ws.cell(r, 5, kt_name).font   = _font(bold=True, size=11)
        ws.cell(r, 5).fill = fill; ws.cell(r, 5).border = bdr
        ws.cell(r, 5).alignment = _center()

        kt_disp = round(kt_val, 4) if isinstance(kt_val, float) else (kt_val if kt_val is not None else "—")
        ktc = ws.cell(r, 6, kt_disp)
        ktc.font = _font(size=11)
        ktc.fill = fill; ktc.border = bdr; ktc.alignment = _right()
        if isinstance(kt_val, float): ktc.number_format = "#,##0.00"

    # Kết luận
    r_kl = 9 + len(rows_sum) + 1
    ws.merge_cells(f"A{r_kl}:F{r_kl}")
    kl_color = (
        _CLR_ACCENT_GREEN if phan_loai == "TỐT"
        else _CLR_ACCENT_ORG if phan_loai == "KHÁ"
        else _CLR_ACCENT_RED
    )
    kl = ws[f"A{r_kl}"]
    kl.value = (
        f"KẾT LUẬN:  Cơ bản {phan_loai} ({tong}/12 điểm)  |  "
        f"Kỹ thuật: {tin_hieu}  |  "
        f"Giải thích: {tech.get('giai_thich', '—')}"
    )
    kl.font = _font(bold=True, size=11, color=_CLR_TITLE_TEXT)
    kl.fill = _fill(kl_color)
    kl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    kl.border = _border_medium()
    ws.row_dimensions[r_kl].height = 40

    for col_letter, width in zip("ABCDEF", [12, 14, 10, 12, 14, 14]):
        ws.column_dimensions[col_letter].width = width


# ---------------------------------------------------------------------------
# Route chính
# ---------------------------------------------------------------------------

@export_bp.route("/xuat-excel/<ma_cp>")
def xuat_excel(ma_cp: str):
    """Xuất báo cáo phân tích cổ phiếu ra file Excel (.xlsx)."""
    if not _HAS_OPENPYXL:
        return jsonify({
            "error": "Thiếu thư viện openpyxl",
            "detail": "Chạy: pip install openpyxl"
        }), 500

    ma_cp = ma_cp.upper()

    try:
        # Thu thập dữ liệu từ 3 module
        info = module1.lay_thong_tin_co_phieu(ma_cp)
    except Exception as e:
        return jsonify({"error": f"Không tìm thấy mã {ma_cp}", "detail": str(e)}), 404

    try:
        df_gia = module1.lay_gia_lich_su(ma_cp)
        tech   = module2.tom_tat_module2(df_gia)
    except Exception as e:
        tech   = {}
        df_gia = None

    try:
        fund = module3.tom_tat_module3(ma_cp)
    except Exception as e:
        fund = {"chi_so": {}, "cham_diem": {}}

    # Tạo Workbook
    wb = openpyxl.Workbook()
    # Xóa sheet mặc định
    wb.remove(wb.active)

    _sheet_tong_hop(wb, info, tech, fund, ma_cp)
    _sheet_thong_tin(wb, info, ma_cp)
    if df_gia is not None and not df_gia.empty:
        _sheet_gia_lich_su(wb, df_gia)
    _sheet_ky_thuat(wb, tech)
    _sheet_co_ban(wb, fund)

    # Xuất ra bộ nhớ
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{ma_cp}_phan_tich_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
